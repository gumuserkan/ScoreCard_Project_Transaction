"""Utility helpers for wallet features."""
from __future__ import annotations

import asyncio
import csv
import logging
import os
import random
import re
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Iterable, List, Optional, Sequence, Set

try:  # pragma: no cover - import guard for optional dependency
    from dateutil import parser
except ImportError:  # pragma: no cover - fallback for test environments
    class _Parser:
        @staticmethod
        def isoparse(value: str) -> datetime:
            value = value.replace("Z", "+00:00")
            return datetime.fromisoformat(value)

    parser = _Parser()  # type: ignore

LOGGER = logging.getLogger(__name__)


WALLET_REGEX = re.compile(r"^0x[a-fA-F0-9]{40}$")


def is_wallet_address(value: str) -> bool:
    """Return True if the given string is a hex Ethereum address."""

    return bool(WALLET_REGEX.fullmatch(value))


def normalize_wallet(value: str) -> str:
    """Normalize ENS or address strings."""

    value = value.strip()
    if not value:
        return value
    if value.startswith("0x"):
        return value.lower()
    return value.lower()


def load_wallets(
    input_path: Optional[Path], wallets_list: Optional[str]
) -> List[str]:
    """Load wallet addresses from CLI parameters."""

    wallets: Set[str] = set()
    if wallets_list:
        for part in wallets_list.split(","):
            part = part.strip()
            if part:
                wallets.add(normalize_wallet(part))
    if input_path:
        if not input_path.exists():
            raise FileNotFoundError(f"Input file {input_path} not found")
        suffix = input_path.suffix.lower()
        if suffix == ".csv":
            with input_path.open("r", newline="", encoding="utf-8") as fh:
                reader = csv.reader(fh)
                for row in reader:
                    if not row:
                        continue
                    wallets.add(normalize_wallet(row[0]))
        elif suffix in {".xlsx", ".xlsm"}:
            try:
                from openpyxl import load_workbook
            except ImportError as exc:  # pragma: no cover - missing optional dependency
                raise ImportError(
                    "Reading Excel files requires the optional dependency 'openpyxl'. "
                    "Install it with `pip install openpyxl`."
                ) from exc
            workbook = load_workbook(filename=input_path, read_only=True, data_only=True)
            try:
                sheet = workbook.active
                header: Optional[List[str]] = None
                address_idx: Optional[int] = None
                for row in sheet.iter_rows(values_only=True):
                    if header is None:
                        header = [str(cell).strip().lower() if cell is not None else "" for cell in row]
                        try:
                            address_idx = header.index("address")
                            continue
                        except ValueError:
                            # If the first row doesn't look like a header, treat it as data.
                            first_non_empty: Optional[int] = None
                            valid_found = False
                            has_values = False
                            for idx, cell in enumerate(row):
                                if cell is None:
                                    continue
                                text = str(cell).strip()
                                if not text:
                                    continue
                                has_values = True
                                if first_non_empty is None:
                                    first_non_empty = idx
                                normalized = normalize_wallet(text)
                                if is_wallet_address(normalized) or normalized.endswith(".eth"):
                                    wallets.add(normalized)
                                    valid_found = True
                            if not has_values:
                                header = None
                                continue
                            if not valid_found:
                                raise ValueError(
                                    f"Excel file {input_path} must contain an 'address' column"
                                )
                            address_idx = first_non_empty if first_non_empty is not None else 0
                            # first row already processed as data
                            continue
                    if address_idx is None or address_idx >= len(row):
                        continue
                    cell = row[address_idx]
                    if cell is None:
                        continue
                    wallets.add(normalize_wallet(str(cell)))
            finally:
                workbook.close()
        else:
            with input_path.open("r", encoding="utf-8") as fh:
                for line in fh:
                    if line.strip():
                        wallets.add(normalize_wallet(line))
    return sorted(wallets)


def load_env_file(path: Optional[Path] = None) -> None:
    """Load environment variables from a .env style file if present."""

    target = path or Path(".env")
    if not target.exists():
        return
    with target.open("r", encoding="utf-8") as handle:
        for raw_line in handle:
            line = raw_line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" not in line:
                continue
            key, _, value = line.partition("=")
            key = key.strip()
            if not key:
                continue
            value = value.strip()
            if (value.startswith('"') and value.endswith('"')) or (
                value.startswith("'") and value.endswith("'")
            ):
                value = value[1:-1]
            os.environ.setdefault(key, value)


def utc_now() -> datetime:
    """Return timezone-aware utc now."""

    return datetime.now(timezone.utc)


def ensure_event_loop() -> asyncio.AbstractEventLoop:
    """Get or create an asyncio event loop."""

    try:
        loop = asyncio.get_running_loop()
    except RuntimeError:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
    return loop


def async_retry(
    retries: int = 5,
    base_delay: float = 0.5,
    jitter: float = 0.2,
    exceptions: Sequence[type[BaseException]] = (Exception,),
):
    """Decorator implementing exponential backoff for async callables."""

    def decorator(func):
        async def wrapper(*args, **kwargs):
            attempt = 0
            while True:
                try:
                    return await func(*args, **kwargs)
                except exceptions as exc:  # pragma: no cover - runtime only
                    attempt += 1
                    if attempt > retries:
                        raise
                    delay = base_delay * (2 ** (attempt - 1))
                    delay += random.random() * jitter
                    LOGGER.warning(
                        "Retrying %s due to %s (attempt %s/%s, sleeping %.2fs)",
                        func.__name__,
                        exc,
                        attempt,
                        retries,
                        delay,
                    )
                    await asyncio.sleep(delay)

        return wrapper

    return decorator


def month_diff(start: datetime, end: datetime) -> int:
    """Return number of calendar months between dates, at least 1."""

    months = (end.year - start.year) * 12 + (end.month - start.month)
    if end.day >= start.day:
        months += 1
    return max(months, 1)


@dataclass(frozen=True)
class TimeWindow:
    """Represents a time window boundary."""

    label: str
    days: int

    def cutoff(self, reference: datetime) -> datetime:
        return reference - timedelta(days=self.days)


def parse_iso8601(value: str) -> datetime:
    """Parse timestamp string to aware datetime."""

    dt = parser.isoparse(value)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


class AsyncLimiter:
    """Simple semaphore-based concurrency limiter for async contexts."""

    def __init__(self, limit: int):
        self._semaphore = asyncio.Semaphore(limit)

    async def __aenter__(self):
        await self._semaphore.acquire()
        return self

    async def __aexit__(self, exc_type, exc, tb):
        self._semaphore.release()
        return False


def getenv(key: str, default: Optional[str] = None) -> Optional[str]:
    """Wrapper around os.getenv for easier mocking."""

    return os.getenv(key, default)


def getenv_bool(key: str, default: bool) -> bool:
    """Return a boolean interpretation of an environment variable."""

    value = getenv(key)
    if value is None:
        return default
    normalized = value.strip().lower()
    if normalized in {"1", "true", "yes", "on"}:
        return True
    if normalized in {"0", "false", "no", "off"}:
        return False
    return default
